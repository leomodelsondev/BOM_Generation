import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
from datetime import datetime
import re

print("Starting BOM Processor v3.0...")

# =====================================================
# READ EXCEL FILE
# =====================================================
def read_excel_file(path):
    """Read Excel file (.xls or .xlsx)"""
    print(f"Reading file: {path}")
    ext = os.path.splitext(path)[1].lower()
    
    if ext == ".xlsx":
        print("Reading as .xlsx format")
        return pd.read_excel(path, engine="openpyxl")
    elif ext == ".xls":
        print("Reading as .xls format")
        import xlwings as xw
        app = xw.App(visible=False)
        wb = app.books.open(path)
        data = wb.sheets[0].used_range.value
        wb.close()
        app.quit()
        return pd.DataFrame(data[1:], columns=data[0])
    else:
        raise ValueError("Unsupported file format. Please use .xls or .xlsx")


# =====================================================
# GENERATE HIERARCHY AND LEVEL FROM PART CODE INDENTATION
# =====================================================
def generate_hierarchy_level(df):
    """
    Generate Hierarchy (1, 1.1, 1.2, etc.) and Level based on Part Code indentation
    Each 2 spaces = 1 level deep
    """
    print("\n" + "-"*40)
    print("GENERATING HIERARCHY & LEVEL...")
    
    # Clean Part Code column
    df['Part Code'] = df['Part Code'].fillna('').astype(str)
    
    # Calculate level based on leading spaces (2 spaces = 1 level)
    df['__spaces'] = df['Part Code'].apply(lambda x: len(x) - len(x.lstrip()))
    df['Level'] = df['__spaces'] // 2  # 2 spaces = 1 level
    
    # Initialize hierarchy counters for each level
    counters = {}  # {level: current_count}
    hierarchy_list = []
    
    for idx, row in df.iterrows():
        current_level = row['Level']
        
        # Reset counters for levels deeper than current
        keys_to_remove = [k for k in counters.keys() if k > current_level]
        for k in keys_to_remove:
            del counters[k]
        
        # Increment counter for current level
        if current_level not in counters:
            counters[current_level] = 1
        else:
            counters[current_level] += 1
        
        # Build hierarchy string
        hierarchy_parts = []
        for level in range(current_level + 1):
            hierarchy_parts.append(str(counters.get(level, 1)))
        
        hierarchy = ".".join(hierarchy_parts)
        hierarchy_list.append(hierarchy)
    
    # Assign to dataframe
    df['Hierarchy'] = hierarchy_list
    
    # Remove temporary column
    df.drop(columns=['__spaces'], inplace=True)
    
    # Count unique levels
    level_counts = df['Level'].value_counts().sort_index()
    print(f"Level distribution:")
    for level, count in level_counts.items():
        print(f"  Level {level}: {count} rows")
    
    print(f"✓ Hierarchy generated for {len(df)} rows")
    
    # Show sample
    print("\nSample (first 15 rows):")
    print(f"{'Row#':<6} {'Part Code':<30} {'Spaces':<8} {'Level':<6} {'Hierarchy'}")
    print("-"*70)
    
    for idx, row in df.head(15).iterrows():
        part_code = str(row['Part Code'])
        stripped = part_code.strip()[:28]
        spaces = len(part_code) - len(part_code.lstrip())
        print(f"{idx+2:<6} '{stripped:<28}' {spaces:<8} {row['Level']:<6} {row['Hierarchy']}")
    
    return df


# =====================================================
# CHECK IF OTHER COLUMNS ARE EMPTY
# =====================================================
def are_other_columns_empty(row, exclude_columns):
    """
    Check if all columns except the excluded ones are empty/NaN/0
    Returns True if all other columns are empty
    """
    for col in row.index:
        if col in exclude_columns:
            continue
        
        value = row[col]
        
        # Skip if NaN
        if pd.isna(value):
            continue
        
        # Skip if empty string
        if str(value).strip() == '':
            continue
        
        # Skip if 0 or 0.0
        try:
            if float(value) == 0:
                continue
        except (ValueError, TypeError):
            pass
        
        # If we get here, column has some data
        return False
    
    return True


# =====================================================
# MAIN PROCESS
# =====================================================
def process_file(file_path, progress_bar, root_window):
    try:
        print("\n" + "="*60)
        print("PROCESS STARTED")
        print("="*60)
        print(f"File: {file_path}")
        
        # Update progress
        progress_bar["value"] = 5
        root_window.update()
        
        # Read the Excel file
        print("Reading Excel file...")
        df = read_excel_file(file_path)
        
        print(f"✓ File loaded successfully")
        print(f"  Total rows: {len(df)}")
        print(f"  Columns: {df.columns.tolist()}")
        
        # Store original row count for summary
        original_row_count = len(df)
        
        progress_bar["value"] = 10
        root_window.update()
        
        # =============================================
        # STEP 0: GENERATE HIERARCHY & LEVEL
        # =============================================
        print("\n" + "="*60)
        print("STEP 0: GENERATING HIERARCHY & LEVEL")
        print("="*60)
        
        df = generate_hierarchy_level(df)
        
        progress_bar["value"] = 20
        root_window.update()
        
        # =============================================
        # CLEAN DATA
        # =============================================
        print("\n" + "-"*40)
        print("CLEANING DATA...")
        
        # Clean Description column
        df['Description'] = df['Description'].fillna('').astype(str).str.strip()
        
        # Clean Material column
        if 'Material' in df.columns:
            df['Material'] = df['Material'].fillna('').astype(str).str.strip()
        
        # Convert numeric columns
        df['Thickness (mm)'] = pd.to_numeric(df['Thickness (mm)'], errors='coerce')
        df['Total Weight (kg)'] = pd.to_numeric(df['Total Weight (kg)'], errors='coerce').fillna(0)
        df['Unit Qty'] = pd.to_numeric(df['Unit Qty'], errors='coerce').fillna(0)
        
        print("✓ Data cleaned")
        
        progress_bar["value"] = 25
        root_window.update()
        
        # =============================================
        # STEP 1: Find all rows with "Sheet" in Description
        # =============================================
        print("\n" + "="*60)
        print("STEP 1: PROCESSING 'SHEET' ROWS")
        print("="*60)
        
        # Find Sheet rows
        sheet_mask = df['Description'].str.lower() == 'sheet'
        sheet_rows = df[sheet_mask]
        
        print(f"✓ Found {len(sheet_rows)} rows with 'Sheet' in Description")
        
        progress_bar["value"] = 35
        root_window.update()
        
        # =============================================
        # STEP 2: Move Thickness from Sheet row to upper row
        # =============================================
        print("\n" + "-"*40)
        print("Moving Thickness to upper row...")
        
        moved_count = 0
        
        for idx, row in sheet_rows.iterrows():
            current_thickness = row['Thickness (mm)']
            
            # Check if thickness is valid (not NaN and > 0)
            if pd.notna(current_thickness) and current_thickness > 0:
                # Get the upper row (idx - 1)
                if idx > 0:
                    upper_idx = idx - 1
                    
                    # Copy thickness to upper row
                    df.at[upper_idx, 'Thickness (mm)'] = current_thickness
                    moved_count += 1
                    print(f"✓ Moved thickness {current_thickness} "
                          f"from Row {idx+2} ({row['Hierarchy']}) "
                          f"to Row {upper_idx+2} ({df.at[upper_idx, 'Hierarchy']})")
                else:
                    print(f"✗ Row {idx+2} is the first row, cannot move thickness up")
            else:
                print(f"- Row {idx+2} ({row['Hierarchy']}): No valid thickness")
        
        print(f"✓ Total thickness values moved: {moved_count}")
        
        progress_bar["value"] = 45
        root_window.update()
        
        # =============================================
        # STEP 3: Delete Sheet rows where Total Weight = 0
        # =============================================
        print("\n" + "="*60)
        print("STEP 2: DELETING 'SHEET' ROWS WITH TOTAL WEIGHT = 0")
        print("="*60)
        
        # Find Sheet rows with Total Weight = 0
        sheet_delete_mask = (df['Description'].str.lower() == 'sheet') & (df['Total Weight (kg)'] == 0)
        sheet_rows_count = sheet_delete_mask.sum()
        
        print(f"✓ Found {sheet_rows_count} Sheet rows with Total Weight = 0")
        
        # Delete these rows
        df = df[~sheet_delete_mask].copy()
        df = df.reset_index(drop=True)
        print(f"✓ Deleted {sheet_rows_count} rows")
        print(f"✓ Remaining rows after Step 2: {len(df)}")
        
        progress_bar["value"] = 55
        root_window.update()
        
        # =============================================
        # STEP 4: Delete empty rows (Part Code is empty/whitespace only)
        # =============================================
        print("\n" + "="*60)
        print("STEP 3: DELETING EMPTY ROWS (No Part Code)")
        print("="*60)
        
        # Find rows where Part Code is empty after stripping
        empty_partcode_mask = df['Part Code'].astype(str).str.strip() == ''
        empty_partcode_count = empty_partcode_mask.sum()
        
        print(f"✓ Found {empty_partcode_count} rows with empty Part Code")
        
        if empty_partcode_count > 0:
            # Show some of the empty rows
            empty_rows = df[empty_partcode_mask]
            print("\nSample empty Part Code rows (first 10):")
            for idx, row in empty_rows.head(10).iterrows():
                desc = str(row['Description'])[:40]
                print(f"  Row {idx+2}: Hierarchy={row['Hierarchy']}, Desc='{desc}'")
            
            # Delete these rows
            df = df[~empty_partcode_mask].copy()
            df = df.reset_index(drop=True)
            print(f"\n✓ Deleted {empty_partcode_count} rows with empty Part Code")
        else:
            print("  No empty Part Code rows found")
        
        print(f"✓ Remaining rows after Step 3: {len(df)}")
        
        progress_bar["value"] = 65
        root_window.update()
        
        # =============================================
        # STEP 5: Delete rows with data in Description/Unit Qty/Material
        #         but Total Weight = 0 and all other columns empty
        # =============================================
        print("\n" + "="*60)
        print("STEP 4: DELETING ROWS WITH PARTIAL DATA & TOTAL WEIGHT = 0")
        print("="*60)
        
        # Define the key columns that should have data
        key_columns = ['Description', 'Unit Qty']
        if 'Material' in df.columns:
            key_columns.append('Material')
        
        # Define columns to exclude when checking if "rest of columns are empty"
        exclude_columns = key_columns + ['Total Weight (kg)']
        
        rows_to_delete_indices = []
        
        for idx, row in df.iterrows():
            # Check if Total Weight is 0
            if row['Total Weight (kg)'] != 0:
                continue
            
            # Check if key columns have any data
            has_key_data = False
            for col in key_columns:
                if col in df.columns:
                    value = row[col]
                    if pd.notna(value) and str(value).strip() != '' and str(value).strip() != '0' and value != 0:
                        has_key_data = True
                        break
            
            if not has_key_data:
                continue
            
            # Check if all other columns are empty
            other_cols_empty = are_other_columns_empty(row, exclude_columns)
            
            if other_cols_empty:
                rows_to_delete_indices.append(idx)
        
        partial_data_count = len(rows_to_delete_indices)
        
        print(f"✓ Found {partial_data_count} rows with partial data & Total Weight = 0")
        
        if partial_data_count > 0:
            # Delete these rows
            df = df.drop(rows_to_delete_indices)
            df = df.reset_index(drop=True)
            print(f"✓ Deleted {partial_data_count} rows")
        else:
            print("  No matching rows found")
        
        print(f"✓ Remaining rows after Step 4: {len(df)}")
        
        progress_bar["value"] = 75
        root_window.update()
        
        # =============================================
        # SUMMARY
        # =============================================
        total_deleted = original_row_count - len(df)
        final_row_count = len(df)
        
        print("\n" + "="*60)
        print("SUMMARY")
        print("="*60)
        print(f"Original rows: {original_row_count}")
        print(f"Sheet thickness moved: {moved_count}")
        print(f"Sheet rows deleted: {sheet_rows_count}")
        print(f"Empty Part Code rows deleted: {empty_partcode_count}")
        print(f"Partial data rows deleted: {partial_data_count}")
        print(f"Total rows deleted: {total_deleted}")
        print(f"Final rows: {final_row_count}")
        
        # Show final hierarchy structure
        print("\nFinal Hierarchy Structure (first 20 rows):")
        print(f"{'Row#':<6} {'Hierarchy':<12} {'Level':<6} {'Part Code'}")
        print("-"*60)
        for idx, row in df.head(20).iterrows():
            print(f"{idx+2:<6} {row['Hierarchy']:<12} {row['Level']:<6} '{row['Part Code'].strip()[:30]}'")
        
        progress_bar["value"] = 85
        root_window.update()
        
        # =============================================
        # REARRANGE COLUMNS - Put Hierarchy and Level first
        # =============================================
        cols = df.columns.tolist()
        
        # Remove Hierarchy and Level if they exist
        for col in ['Hierarchy', 'Level']:
            if col in cols:
                cols.remove(col)
        
        # Add them at the beginning
        final_cols = ['Hierarchy', 'Level'] + cols
        df = df[final_cols]
        
        # =============================================
        # SAVE OUTPUT
        # =============================================
        print("\n" + "-"*40)
        print("SAVING OUTPUT FILE...")
        
        # Create output filename
        base_name = os.path.splitext(file_path)[0]
        output_path = f"{base_name}_processed.xlsx"
        
        print(f"Output path: {output_path}")
        
        # Save with openpyxl
        print("Writing to Excel file...")
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write data starting from row 7
            df.to_excel(writer, index=False, startrow=6, sheet_name='BOM')
            
            # Add header information
            worksheet = writer.sheets['BOM']
            
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, size=11, color="2F5496")
            value_font = Font(size=11)
            title_font = Font(bold=True, size=14, color="1F3864")
            
            # Row 1 - Title (DO NOT merge cells - causes issues)
            worksheet['A1'] = "BOM PROCESSING REPORT"
            worksheet['A1'].font = title_font
            
            # Row 2
            worksheet['A2'] = "Processed Date:"
            worksheet['A2'].font = header_font
            worksheet['B2'] = datetime.today().strftime("%d-%m-%Y")
            worksheet['B2'].font = value_font
            
            # Row 3
            worksheet['A3'] = "Original File:"
            worksheet['A3'].font = header_font
            worksheet['B3'] = os.path.basename(file_path)
            worksheet['B3'].font = value_font
            
            # Row 4
            worksheet['A4'] = "Original Rows:"
            worksheet['A4'].font = header_font
            worksheet['B4'] = original_row_count
            worksheet['B4'].font = value_font
            
            worksheet['C4'] = "Final Rows:"
            worksheet['C4'].font = header_font
            worksheet['D4'] = final_row_count
            worksheet['D4'].font = value_font
            
            # Row 5
            worksheet['A5'] = "Rows Deleted:"
            worksheet['A5'].font = header_font
            worksheet['B5'] = total_deleted
            worksheet['B5'].font = value_font
            
            worksheet['C5'] = "Thickness Moved:"
            worksheet['C5'].font = header_font
            worksheet['D5'] = moved_count
            worksheet['D5'].font = value_font
            
            # Row 6 is empty (separator)
            
            # Style the header row (row 7 - the actual column headers)
            header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            header_font_white = Font(bold=True, size=10, color="FFFFFF")
            
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=7, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font_white
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adjust column widths safely (skip merged cells)
            for col_idx in range(1, len(df.columns) + 1):
                max_length = 0
                column_letter = worksheet.cell(row=1, column=col_idx).column_letter
                
                for row_idx in range(1, len(df) + 8):  # +8 for header rows
                    try:
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Set specific widths for Hierarchy and Level
            worksheet.column_dimensions['A'].width = 12  # Hierarchy
            worksheet.column_dimensions['B'].width = 8   # Level
        
        progress_bar["value"] = 100
        root_window.update()
        
        print(f"✓ Output saved to: {output_path}")
        print("\n" + "="*60)
        print("PROCESS COMPLETED SUCCESSFULLY!")
        print("="*60)
        
        # Show success message
        messagebox.showinfo(
            "SUCCESS",
            f"Process completed successfully!\n\n"
            f"📊 Summary:\n"
            f"• Original rows: {original_row_count}\n"
            f"• Sheet thickness moved: {moved_count}\n"
            f"• Sheet rows deleted: {sheet_rows_count}\n"
            f"• Empty Part Code rows deleted: {empty_partcode_count}\n"
            f"• Partial data rows deleted: {partial_data_count}\n"
            f"• Total rows deleted: {total_deleted}\n"
            f"• Final rows: {final_row_count}\n\n"
            f"💾 Output saved as:\n{os.path.basename(output_path)}"
        )
        
    except Exception as e:
        progress_bar["value"] = 0
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        messagebox.showerror("ERROR", f"An error occurred:\n\n{str(e)}")


# =====================================================
# GUI
# =====================================================
def create_gui():
    """Create and run the GUI"""
    print("Creating GUI window...")
    
    root = tk.Tk()
    root.title("BOM Processor v3.0")
    root.geometry("620x520")
    
    root.lift()
    root.attributes('-topmost', True)
    root.after(1000, lambda: root.attributes('-topmost', False))
    
    # Title
    tk.Label(
        root, 
        text="BOM Processor with Hierarchy Generation",
        font=("Arial", 14, "bold"),
        fg="#1F3864"
    ).pack(pady=15)
    
    # Description
    desc_frame = tk.LabelFrame(root, text="What this tool does", padx=10, pady=10)
    desc_frame.pack(fill=tk.X, padx=20, pady=5)
    
    steps = [
        "1. Generates Hierarchy (1, 1.1, 1.2...) based on Part Code indentation",
        "2. Moves 'Thickness' from 'Sheet' rows to the row above",
        "3. Deletes 'Sheet' rows where Total Weight = 0",
        "4. Deletes rows with empty Part Code",
        "5. Deletes rows with partial data where Total Weight = 0"
    ]
    
    for step in steps:
        tk.Label(desc_frame, text=step, font=("Arial", 9), fg="#333333").pack(anchor='w', pady=2)
    
    # File selection
    file_frame = tk.Frame(root)
    file_frame.pack(pady=15)
    
    tk.Label(file_frame, text="Select Excel File:", font=("Arial", 10)).grid(row=0, column=0, padx=5)
    
    file_entry = tk.Entry(file_frame, width=45, font=("Arial", 9))
    file_entry.grid(row=0, column=1, padx=5)
    
    def browse_file():
        filename = filedialog.askopenfilename(
            title="Select BOM Excel File",
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")]
        )
        if filename:
            file_entry.delete(0, tk.END)
            file_entry.insert(0, filename)
    
    tk.Button(
        file_frame, text="Browse", command=browse_file,
        bg="#FF9800", fg="white", font=("Arial", 9, "bold"),
        cursor="hand2", width=10
    ).grid(row=0, column=2, padx=5)
    
    # Progress bar
    progress_frame = tk.Frame(root)
    progress_frame.pack(pady=15)
    
    progress_bar = ttk.Progressbar(progress_frame, length=450, mode='determinate')
    progress_bar.pack()
    
    # Status label
    status_label = tk.Label(root, text="", fg="blue", font=("Arial", 9))
    status_label.pack(pady=5)
    
    # Process button
    def start_processing():
        file_path = file_entry.get().strip()
        
        if not file_path:
            messagebox.showerror("Error", "Please select an Excel file first!")
            return
        
        if not os.path.exists(file_path):
            messagebox.showerror("Error", f"File does not exist:\n{file_path}")
            return
        
        status_label.config(text="Processing... Please wait")
        progress_bar["value"] = 0
        root.update()
        
        thread = threading.Thread(
            target=process_file,
            args=(file_path, progress_bar, root),
            daemon=True
        )
        thread.start()
        
        def check_done():
            if progress_bar["value"] >= 100:
                status_label.config(text="✅ Completed!")
            else:
                root.after(500, check_done)
        
        root.after(500, check_done)
    
    tk.Button(
        root, text="▶ START PROCESSING", command=start_processing,
        bg="#4CAF50", fg="white", font=("Arial", 11, "bold"),
        padx=30, pady=10, cursor="hand2"
    ).pack(pady=10)
    
    tk.Label(
        root, text="Check terminal for detailed progress | BOM Processor v3.0",
        fg="gray", font=("Arial", 8)
    ).pack(side=tk.BOTTOM, pady=10)
    
    root.mainloop()


# =====================================================
# MAIN
# =====================================================
if __name__ == "__main__":
    print("\n" + "="*60)
    print("INITIALIZING BOM PROCESSOR v3.0")
    print("="*60)
    create_gui()