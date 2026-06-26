import os, re, shutil
import pandas as pd
import xlrd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# =====================================================
# DEFAULT TEMPLATE PATH
# =====================================================
TEMPLATE_PATH = r'D:\MAC Excel Code\VS Code\BOM_Generation\Template-Standard_BOM.xlsx'

FILL_GREEN  = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
FILL_NONE   = PatternFill(fill_type=None)
FONT_BOLD   = Font(name="Calibri", size=11, bold=True)
FONT_NORMAL = Font(name="Calibri", size=11, bold=False)


# =====================================================
# READ XLS/XLSX — rebuild Hierarchy from row order
# to avoid Excel float problem (1.10 read as 1.1)
# =====================================================
def read_and_fix_hierarchy(path):
    ext = os.path.splitext(path)[1].lower()

    if ext == ".xls":
        wb  = xlrd.open_workbook(path)
        ws  = wb.sheets()[0]
        headers = [str(ws.cell(0, c).value).strip() for c in range(ws.ncols)]
        rows = []
        for ri in range(1, ws.nrows):
            row = {}
            for ci, h in enumerate(headers):
                row[h] = ws.cell(ri, ci).value
            rows.append(row)
        df = pd.DataFrame(rows)
    else:
        df = pd.read_excel(path, engine="openpyxl", dtype=str)

    # Strip all string columns
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].fillna('').astype(str).str.strip()

    # ── REBUILD HIERARCHY FROM ROW ORDER ────────────────
    # Strategy: parse the raw Hierarchy string segments,
    # but track counters per depth so 1.10 ≠ 1.1
    # The raw Hierarchy from XLS may have 1.0 (=level 1),
    # 1.1, 1.2 ... 1.9, 1.1 (=1.10), 1.11 etc.
    # We walk rows in order, parse depth from segment count,
    # and increment counter at each depth independently.

    raw_hierarchies = df['Hierarchy'].astype(str).tolist()

    counters   = {}   # depth -> current count
    new_hier   = []
    prev_depth = 0

    for raw in raw_hierarchies:
        raw = raw.strip().rstrip('.0').strip()
        # Count depth = number of dots + 1
        # But for root "1" or "1.0" depth = 1
        parts = [p for p in raw.split('.') if p != '']
        depth = len(parts)
        if depth == 0:
            depth = 1

        # If depth increased, reset all deeper counters
        if depth > prev_depth:
            for d in range(prev_depth + 1, depth + 1):
                if d not in counters:
                    counters[d] = 0
            counters[depth] = counters.get(depth, 0) + 1
        elif depth == prev_depth:
            counters[depth] = counters.get(depth, 0) + 1
        else:
            # depth decreased — reset deeper counters
            for d in list(counters.keys()):
                if d > depth:
                    del counters[d]
            counters[depth] = counters.get(depth, 0) + 1

        hier = '.'.join(str(counters.get(d, 1)) for d in range(1, depth + 1))
        new_hier.append(hier)
        prev_depth = depth

    df['Hierarchy'] = new_hier

    print("\n  Hierarchy rebuild (first 35 rows):")
    for i, (raw, new) in enumerate(zip(raw_hierarchies, new_hier)):
        print(f"    row {i+1:2d}: raw='{raw}'  ->  rebuilt='{new}'")

    return df


# =====================================================
# EXTRACT THICKNESS FROM PART CODE
# =====================================================
def extract_thickness(part_code):
    code = str(part_code).strip().upper()
    m = re.search(r'-T(\d+\.?\d*)(?:-|$)', code)
    if not m:
        m = re.search(r'T(\d+\.?\d*)', code)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            pass
    return None


# =====================================================
# VALID CHILD PART CHECK
# Part code must end with -<digits only> (2-7 digits)
# =====================================================
def is_valid_child_part(part_code):
    pc = str(part_code).strip()
    return bool(re.search(r'-\d{2,7}$', pc))


# =====================================================
# LEAF / PARENT TAGGING
# =====================================================
def tag_leaf_parent(df):
    hierarchies = df['Hierarchy'].astype(str).tolist()
    def is_leaf(h):
        prefix = str(h) + '.'
        return not any(str(o).startswith(prefix) for o in hierarchies if str(o) != str(h))
    df = df.copy()
    df['__is_leaf'] = df['Hierarchy'].apply(is_leaf)
    return df


# =====================================================
# DIRECT CHILDREN
# =====================================================
def get_direct_children(parent_h, df):
    parent_h = str(parent_h)
    prefix   = parent_h + '.'
    result   = []
    for _, row in df.iterrows():
        h = str(row['Hierarchy'])
        if h.startswith(prefix) and '.' not in h[len(prefix):]:
            result.append(row)
    return result


# =====================================================
# BUILD RM LOOKUP
# =====================================================
def build_rm_lookup(template_path):
    rm_df = pd.read_excel(template_path, sheet_name='Raw Material Part Code')
    fl_lookup  = {}
    be_options = []
    for _, row in rm_df.iterrows():
        ic    = str(row['Item Code']).strip()
        iname = str(row['Item Name']).strip()
        m = re.search(r'-(\d+\.?\d*)$', ic)
        if m:
            try:
                fl_lookup[float(m.group(1))] = (ic, iname)
            except ValueError:
                pass
        if 'sheet' in iname.lower():
            be_options.append((ic, iname))
    print(f"  FL lookup entries : {len(fl_lookup)}")
    print(f"  BE sheet options  : {len(be_options)}")
    return fl_lookup, be_options


# =====================================================
# WRITE CHILD PART BOM SHEET
# =====================================================
def write_child_sheet(ws, leaf_rows):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
            cell.fill  = FILL_NONE

    for i, r in enumerate(leaf_rows, start=2):
        ws.cell(row=i, column=1, value=r['item_code']).font   = FONT_NORMAL
        ws.cell(row=i, column=2, value=r['item_name']).font   = FONT_NORMAL
        ws.cell(row=i, column=3, value=r['description']).font = FONT_NORMAL
        for col, val in [(4, r['uom']), (5, r['routing']),
                         (6, r['qty']), (7, r['weight'])]:
            c = ws.cell(row=i, column=col, value=val)
            c.font      = FONT_NORMAL
            c.alignment = Alignment(horizontal='center')
        # H: yellow fill only when auto-filled
        ch = ws.cell(row=i, column=8, value=r['item_bom'])
        ch.font = FONT_NORMAL
        ch.fill = FILL_YELLOW if r['item_bom'] and not str(r['item_bom']).startswith('[') else FILL_NONE
        for col, val in [(9,  r['item_name_bom']),
                         (10, r['qty_bom']),
                         (11, r['uom_bom']),
                         (12, r['do_not_explode'])]:
            c = ws.cell(row=i, column=col, value=val)
            c.font      = FONT_NORMAL
            c.alignment = Alignment(horizontal='center')

    print(f"  Written {len(leaf_rows)} rows to Child Part BOM")


# =====================================================
# WRITE PARENT BOM SHEET
# =====================================================
def write_parent_sheet(ws, parent_sections):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
            cell.fill  = FILL_NONE
            cell.font  = FONT_NORMAL

    current_row = 2
    for section in parent_sections:
        p = section['parent']
        ca = ws.cell(row=current_row, column=1, value=p['item_code'])
        ca.font = FONT_BOLD
        ca.fill = FILL_GREEN
        ws.cell(row=current_row, column=2, value=p['item_name']).alignment  = Alignment(horizontal='center')
        ws.cell(row=current_row, column=3, value=p['description']).alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=4, value=p['uom']).alignment         = Alignment(horizontal='center')
        ws.cell(row=current_row, column=5, value=p['routing']).alignment     = Alignment(horizontal='center')
        cf = ws.cell(row=current_row, column=6, value=1)
        cf.alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=7, value=p['weight']).alignment = Alignment(horizontal='center')
        current_row += 1

        for child in section['children']:
            ch = ws.cell(row=current_row, column=8, value=child['item_code'])
            ch.fill = FILL_YELLOW
            ws.cell(row=current_row, column=9,  value=child['item_name']).alignment  = Alignment(horizontal='center')
            ws.cell(row=current_row, column=10, value=child['qty']).alignment        = Alignment(horizontal='center')
            ws.cell(row=current_row, column=11, value=child['uom']).alignment        = Alignment(horizontal='center')
            ws.cell(row=current_row, column=12, value=child['do_not_explode']).alignment = Alignment(horizontal='center')
            current_row += 1

    print(f"  Written {current_row - 2} rows to Parent BOM")
    return current_row - 2


# =====================================================
# MAIN PROCESS
# =====================================================
def process_file(file_path, template_path, progress_bar, root_window, status_label):
    try:
        def upd(msg, pct):
            status_label.config(text=msg)
            progress_bar["value"] = pct
            root_window.update()

        print("\n" + "="*60)
        print("  ERP BOM FORMATTER  v4.0")
        print("="*60)

        # ── READ & FIX HIERARCHY ─────────────────────────────
        upd("Reading & rebuilding hierarchy ...", 5)
        df_raw = read_and_fix_hierarchy(file_path)
        original_count = len(df_raw)
        print(f"\n  Rows loaded : {original_count}")

        # Identify PART NUMBER column (may be 'PART NUMBER' or 'Part Code')
        pc_col = 'PART NUMBER' if 'PART NUMBER' in df_raw.columns else 'Part Code'
        print(f"  Part Code column: '{pc_col}'")

        df_raw[pc_col]        = df_raw[pc_col].fillna('').astype(str).str.strip()
        df_raw['Description'] = df_raw['Description'].fillna('').astype(str).str.strip()
        if 'Part Name' in df_raw.columns:
            df_raw['Part Name'] = df_raw['Part Name'].fillna('').astype(str).str.strip()

        # ── LOAD RM LOOKUP ───────────────────────────────────
        upd("Loading RM reference ...", 12)
        fl_lookup, be_options = build_rm_lookup(template_path)

        # ── STEP 1: DELETE UNWANTED ROWS ─────────────────────
        upd("Step 1: Deleting blank Part Code rows ...", 20)
        print("\nSTEP 1 - DELETE BLANK Part Code ROWS")
        is_blank  = df_raw[pc_col] == ''
        rule_a    = is_blank & (df_raw['Description'].str.lower() == 'sheet')
        rule_b    = is_blank & (df_raw['Description'].str.lower() != 'sheet')
        to_del    = rule_a | rule_b
        del_count = int(to_del.sum())
        print(f"  Rule A (Sheet + blank) : {rule_a.sum()}")
        print(f"  Rule B (Other + blank) : {rule_b.sum()}")
        print(f"  Total deleted          : {del_count}")
        df = df_raw[~to_del].reset_index(drop=True)
        print(f"  Rows remaining         : {len(df)}")

        # ── STEP 2: TAG LEAF / PARENT ────────────────────────
        upd("Step 2: Tagging hierarchy ...", 30)
        df = tag_leaf_parent(df)

        # ── STEP 3: BUILD CHILD PART BOM ─────────────────────
        upd("Step 3: Building Child Part BOM ...", 45)
        print("\nSTEP 3 - CHILD PART BOM (leaf parts, valid part codes only, no duplicates)")

        leaf_rows        = []
        seen_child_codes = set()   # for duplicate detection
        be_manual_needed = []
        skipped_hw       = []

        for _, row in df[df['__is_leaf']].iterrows():
            pc    = str(row[pc_col]).strip()
            pname = str(row.get('Part Name', '')).strip() or str(row.get('Description', '')).strip()
            desc  = str(row.get('Description', '')).strip()

            # ── FILTER: only valid child parts (end with -<digits>) ──
            if not is_valid_child_part(pc):
                skipped_hw.append(pc)
                print(f"  SKIP (HW/BO): '{pc}'")
                continue

            # ── FILTER: skip duplicates ──────────────────────
            if pc in seen_child_codes:
                print(f"  SKIP (dup): '{pc}'")
                continue
            seen_child_codes.add(pc)

            try:    wt = float(row.get('Unit Weight (kg)', row.get('Unit Weight (Kgs)', 0)) or 0)
            except: wt = ''

            item_bom, item_name_bom = '', ''
            pc_up = pc.upper()

            if pc_up.startswith('FL'):
                t = extract_thickness(pc)
                if t is not None and t in fl_lookup:
                    item_bom, item_name_bom = fl_lookup[t]
                    print(f"  FL auto-fill: {pc} -> T{t} -> {item_bom}")
                else:
                    item_bom      = f'[LOOKUP FAILED T={t}]'
                    item_name_bom = '[MANUAL REQUIRED]'
                    print(f"  FL FAILED: {pc} T={t}")
            elif pc_up.startswith('BE'):
                item_bom      = '[SELECT FROM DROPDOWN]'
                item_name_bom = '[SELECT FROM DROPDOWN]'
                be_manual_needed.append(pc)

            leaf_rows.append({
                'item_code'      : pc,
                'item_name'      : pname,
                'description'    : desc,
                'uom'            : 'Nos',
                'routing'        : '',
                'qty'            : 1,       # Point 3: always 1
                'weight'         : wt,
                'item_bom'       : item_bom,
                'item_name_bom'  : item_name_bom,
                'qty_bom'        : wt,      # same as unit weight
                'uom_bom'        : 'Kg',
                'do_not_explode' : 1,
            })

        print(f"  Valid leaf rows added  : {len(leaf_rows)}")
        print(f"  HW/BO parts skipped   : {len(skipped_hw)}")
        if be_manual_needed:
            print(f"  BE parts (manual RM)  : {len(be_manual_needed)}")

        # ── STEP 4: BUILD PARENT BOM ─────────────────────────
        upd("Step 4: Building Parent BOM ...", 62)
        print("\nSTEP 4 - PARENT BOM (no duplicate parent codes)")

        parent_sections  = []
        seen_parent_codes = set()   # for duplicate parent detection

        for _, prow in df[~df['__is_leaf']].iterrows():
            pc    = str(prow[pc_col]).strip()
            pname = str(prow.get('Part Name', '')).strip() or str(prow.get('Description', '')).strip()
            desc  = str(prow.get('Description', '')).strip()

            # ── FILTER: skip duplicate parents ───────────────
            if pc in seen_parent_codes:
                print(f"  SKIP parent (dup): '{pc}'")
                continue
            seen_parent_codes.add(pc)

            try:    wt = float(prow.get('Unit Weight (kg)', prow.get('Unit Weight (Kgs)', 0)) or 0)
            except: wt = ''

            children_out = []
            for child in get_direct_children(prow['Hierarchy'], df):
                cpc   = str(child[pc_col]).strip()
                cname = str(child.get('Part Name', '')).strip() or str(child.get('Description', '')).strip()
                try:    cqty = float(child.get('Unit Qty', 1) or 1)
                except: cqty = 1.0
                children_out.append({
                    'item_code'     : cpc,
                    'item_name'     : cname,
                    'qty'           : cqty,
                    'uom'           : 'Nos',
                    'do_not_explode': 1,
                })

            parent_sections.append({
                'parent': {
                    'item_code'  : pc,
                    'item_name'  : pname,
                    'description': desc,
                    'uom'        : 'Nos',
                    'routing'    : '',
                    'weight'     : wt,
                },
                'children': children_out,
            })

        print(f"  Parent sections : {len(parent_sections)}")

        # ── SAVE INTO TEMPLATE COPY ───────────────────────────
        upd("Saving into template ...", 78)
        base        = os.path.splitext(file_path)[0]
        output_path = f"{base}_ERP_BOM.xlsx"

        shutil.copy2(template_path, output_path)
        wb = load_workbook(output_path)

        write_child_sheet(wb['Child Part Bom With RM'], leaf_rows)
        write_parent_sheet(wb['Bom Template-1'], parent_sections)

        # BE reference sheet
        if be_options:
            if 'BE_RM_Reference' in wb.sheetnames:
                del wb['BE_RM_Reference']
            ws_ref = wb.create_sheet('BE_RM_Reference')
            ws_ref['A1'] = 'BE Part RM Options -- Item Name contains Sheet'
            ws_ref['A1'].font = Font(bold=True, size=11, name="Calibri")
            ws_ref['A2'] = 'Item Code'
            ws_ref['B2'] = 'Item Name'
            ws_ref['A2'].font = ws_ref['B2'].font = Font(bold=True, name="Calibri")
            for ri, (ic, iname) in enumerate(be_options, start=3):
                ws_ref.cell(row=ri, column=1, value=ic)
                ws_ref.cell(row=ri, column=2, value=iname)
            ws_ref.column_dimensions['A'].width = 32
            ws_ref.column_dimensions['B'].width = 52

        wb.save(output_path)
        upd("Done!", 100)

        print(f"\n  Saved : {output_path}")
        print("="*60)

        be_note = ""
        if be_manual_needed:
            be_note = (f"\n\nNOTE: {len(be_manual_needed)} BE parts need manual RM selection."
                       f"\nSee 'BE_RM_Reference' sheet for options.")

        messagebox.showinfo(
            "Success",
            f"ERP BOM generated!\n\n"
            f"Summary\n"
            f"{'-'*36}\n"
            f"  Input rows           : {original_count}\n"
            f"  Blank PC deleted     : {del_count}\n"
            f"  HW/BO parts skipped  : {len(skipped_hw)}\n"
            f"  Child BOM rows       : {len(leaf_rows)}\n"
            f"  Parent BOM sections  : {len(parent_sections)}\n\n"
            f"Saved as:\n  {os.path.basename(output_path)}"
            + be_note
        )

    except Exception as e:
        progress_bar["value"] = 0
        status_label.config(text="Error -- check terminal")
        print(f"\nERROR: {e}")
        import traceback; traceback.print_exc()
        messagebox.showerror("Error", f"An error occurred:\n\n{e}")


# =====================================================
# GUI
# =====================================================
def create_gui():
    NAVY = "#1B2A4A"; TEAL = "#1E6E6E"; ACCENT = "#E8A838"
    BG   = "#F0F4F8"; WHITE = "#FFFFFF"; GREY = "#7A8A99"; DARK = "#2C3E50"

    root = tk.Tk()
    root.title("ERP BOM Formatter  v4.0")
    root.configure(bg=BG)
    root.geometry("740x580")
    root.resizable(False, False)
    root.lift()
    root.focus_force()

    hdr = tk.Frame(root, bg=NAVY, height=68)
    hdr.pack(fill=tk.X)
    hdr.pack_propagate(False)
    tk.Label(hdr, text="ERP BOM Formatter", font=("Montserrat", 18, "bold"),
             bg=NAVY, fg=WHITE).pack(side=tk.LEFT, padx=22, pady=14)
    tk.Label(hdr, text="v4.0", font=("Montserrat", 10),
             bg=NAVY, fg=ACCENT).pack(side=tk.LEFT, pady=20)

    ftr = tk.Frame(root, bg=NAVY, height=28)
    ftr.pack(fill=tk.X, side=tk.BOTTOM)
    ftr.pack_propagate(False)
    tk.Label(ftr,
             text="Concept & Developed by Mahesh Arvind Chavan  |  ERP BOM Formatter v4.0  |  Nido Automation",
             bg=NAVY, fg=GREY, font=("Montserrat", 8)).pack(pady=6)

    body = tk.Frame(root, bg=BG)
    body.pack(fill=tk.BOTH, expand=True, padx=22, pady=14)

    sf = tk.LabelFrame(body, text="  Processing Steps  ", bg=BG, fg=TEAL,
                       font=("Montserrat", 9, "bold"), bd=1, relief=tk.GROOVE, padx=8, pady=6)
    sf.pack(fill=tk.X, pady=(0, 10))
    for num, text in [
        ("1", "Rebuild Hierarchy from row order -- fixes Excel float issue (1.10 vs 1.1)"),
        ("2", "Delete rows where Part Code is blank"),
        ("3", "Child BOM: only valid parts ending with -<digits>; no HW/BO; no duplicates"),
        ("4", "FL parts: Item (BOM Details) auto-filled from thickness; BE: marked for manual"),
        ("5", "Parent BOM: parent (green) + direct children (yellow); no duplicate parents"),
        (" ", "Output written directly into template copy -- exact same format preserved"),
    ]:
        rf = tk.Frame(sf, bg=BG)
        rf.pack(fill=tk.X, pady=1)
        lb, lf = (TEAL, WHITE) if num.strip().isdigit() else (BG, BG)
        tk.Label(rf, text=f" {num} ", bg=lb, fg=lf,
                 font=("Montserrat", 8, "bold")).pack(side=tk.LEFT, padx=(0, 7))
        tk.Label(rf, text=text, bg=BG, fg=DARK,
                 font=("Montserrat", 9), anchor='w').pack(side=tk.LEFT)

    def make_file_row(parent, label, default='', filetypes=None):
        lf = tk.LabelFrame(parent, text=f"  {label}  ", bg=BG, fg=TEAL,
                            font=("Montserrat", 9, "bold"), bd=1, relief=tk.GROOVE, padx=10, pady=8)
        lf.pack(fill=tk.X, pady=(0, 8))
        fi = tk.Frame(lf, bg=BG)
        fi.pack(fill=tk.X)
        entry = tk.Entry(fi, width=66, font=("Montserrat", 9), bg=WHITE, relief=tk.FLAT,
                         highlightthickness=1, highlightbackground="#CCCCCC")
        if default:
            entry.insert(0, default)
        entry.pack(side=tk.LEFT, padx=(0, 8), ipady=5)
        ft = filetypes or [("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")]
        def browse(e=entry, ft=ft):
            fn = filedialog.askopenfilename(title=f"Select {label}", filetypes=ft)
            if fn:
                e.delete(0, tk.END)
                e.insert(0, fn)
        btn = tk.Button(fi, text="Browse...", command=browse, bg=ACCENT, fg=NAVY,
                        font=("Montserrat", 9, "bold"), relief=tk.FLAT, cursor="hand2",
                        padx=10, pady=5)
        btn.pack(side=tk.LEFT)
        btn.bind("<Enter>", lambda e, b=btn: b.config(bg="#D4932A"))
        btn.bind("<Leave>", lambda e, b=btn: b.config(bg=ACCENT))
        return entry

    tmpl_entry = make_file_row(body, "Template File (.xlsx)",
                               default=TEMPLATE_PATH,
                               filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")])
    file_entry = make_file_row(body, "Input BOM File (.xls / .xlsx)")

    pgf = tk.Frame(body, bg=BG)
    pgf.pack(fill=tk.X, pady=(6, 0))
    style = ttk.Style()
    style.theme_use('clam')
    style.configure("ERP.Horizontal.TProgressbar",
                    troughcolor="#D0D8E4", background=TEAL, thickness=14)
    progress_bar = ttk.Progressbar(pgf, length=696, mode='determinate',
                                   style="ERP.Horizontal.TProgressbar")
    progress_bar.pack(pady=(6, 3))

    status_lbl = tk.Label(body, text="Ready -- select files and press START",
                          bg=BG, fg=GREY, font=("Montserrat", 9, "italic"))
    status_lbl.pack(pady=(3, 8))

    def start():
        fp = file_entry.get().strip()
        tp = tmpl_entry.get().strip()
        if not fp:
            messagebox.showerror("Missing File", "Please select the Input BOM file.")
            return
        if not tp:
            messagebox.showerror("Missing Template", "Please select the Template file.")
            return
        for p, label in [(fp, "Input BOM"), (tp, "Template")]:
            if not os.path.exists(p):
                messagebox.showerror("Not Found", f"{label} file not found:\n{p}")
                return
        progress_bar["value"] = 0
        status_lbl.config(text="Processing ...", fg=TEAL)
        root.update()
        threading.Thread(target=process_file,
                         args=(fp, tp, progress_bar, root, status_lbl),
                         daemon=True).start()

    btn_row = tk.Frame(body, bg=BG)
    btn_row.pack(pady=2)
    start_btn = tk.Button(btn_row, text="START PROCESSING", command=start,
                          bg=TEAL, fg=WHITE, font=("Montserrat", 12, "bold"),
                          relief=tk.FLAT, cursor="hand2", padx=44, pady=10)
    start_btn.pack()
    start_btn.bind("<Enter>", lambda e: start_btn.config(bg=NAVY))
    start_btn.bind("<Leave>", lambda e: start_btn.config(bg=TEAL))

    root.mainloop()


if __name__ == "__main__":
    print("\n" + "="*60)
    print("  INITIALIZING ERP BOM FORMATTER  v4.0")
    print("="*60)
    create_gui()