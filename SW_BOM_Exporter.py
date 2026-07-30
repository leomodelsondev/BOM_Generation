"""
SW_BOM_Exporter.py  v3.0
Exports Indented BOM from a SolidWorks 2021 assembly via Win32COM API.
Columns match Nido_Mfg_BOM_R3.sldbomtbt template exactly.
v3.0 -- batch property read + mass caching for maximum speed.
Developer: Mahesh Arvind Chavan | NIDO Automation
"""

import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
from datetime import datetime
import win32com.client as win32
import pythoncom


SW_PART_TYPE_ASSEMBLY = 2

CP_PART_CODE   = "Part Code"
CP_PART_NAME   = "Part Name"
CP_DESCRIPTION = "Description"
CP_REVISION    = "Revision"
CP_MATERIAL    = "Material"
CP_THICKNESS   = "Thickness (mm)"
CP_PROCESS     = "Process"
CP_PROCESS1    = "Process 1"
CP_PROCESS2    = "Process 2"
CP_REQUIRED    = "Required"
CP_REMARK      = "Remark"

# Module-level caches -- cleared at the start of every export
_prop_cache = {}   # (path::cfg) -> {prop_name_lower: resolved_value}
_mass_cache = {}   # (path::cfg) -> mass_kg float


# =====================================================
# SOLIDWORKS CONNECTION
# =====================================================
def connect_to_solidworks():
    try:
        print("  Connecting to running SolidWorks ...")
        sw = win32.GetActiveObject("SldWorks.Application")
        sw.Visible = True
        print(f"  Connected. Revision: {sw.RevisionNumber}")
        return sw, None
    except Exception as e:
        print(f"  SW not running: {e}")
        return None, str(e)


def launch_and_open(file_path):
    try:
        print(f"  Launching SolidWorks: {file_path}")
        sw = win32.Dispatch("SldWorks.Application")
        sw.Visible = True
        errors   = win32.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
        warnings = win32.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
        doc = sw.OpenDoc6(file_path, SW_PART_TYPE_ASSEMBLY, 1, "", errors, warnings)
        if doc is None:
            return sw, None, f"OpenDoc6 failed. Error: {errors.value}"
        return sw, doc, None
    except Exception as e:
        return None, None, str(e)


def get_active_assembly(sw):
    try:
        doc = sw.ActiveDoc
        if doc is None:
            return None, "No active document in SolidWorks."
        if doc.GetType != SW_PART_TYPE_ASSEMBLY:
            return None, "Active document is not an Assembly."
        print(f"  Active assembly: {doc.GetTitle}")
        return doc, None
    except Exception as e:
        return None, str(e)


# =====================================================
# EXPRESSION DETECTION + SW SYSTEM PROPERTY RESOLVER
# =====================================================
def _is_expr(val):
    """True if val is an unresolved SW link/formula expression."""
    if not val:
        return False
    s = str(val).strip()
    return "@" in s or s.startswith("$PRP") or s.startswith("SW-")


def _get_material_name(model):
    """
    Get material name from SW model using multiple methods in priority order.
    MaterialIdName returns internal DB index for custom materials -- unreliable.
    The correct approach is GetMaterialPropertyName2 which returns the display name.
    Falls back through multiple APIs until a clean text name is found.
    """
    # Method 1: GetMaterialPropertyName2 -- returns display name e.g. "Plain Carbon Steel"
    # This is the same value shown in the SW material dialog title
    try:
        cfg = model.ConfigurationManager.ActiveConfiguration.Name
        mat = model.GetMaterialPropertyName2(cfg, None)
        if mat and str(mat).strip() and not str(mat).strip().isdigit():
            return str(mat).strip()
    except:
        pass

    # Method 2: Read SW-Material via CustomPropertyManager with special blank config
    # SW stores SW-Material as a virtual property readable this way in 2021+
    try:
        cpm = model.Extension.CustomPropertyManager("")
        if cpm:
            res = cpm.Get5("SW-Material", False, "", "", False)
            if isinstance(res, (list, tuple)) and len(res) >= 4:
                ve = str(res[3] or "").strip()
                vo = str(res[2] or "").strip()
                # Use whichever is a readable material name (not an expression, not a number)
                for v in [ve, vo]:
                    if v and not v.isdigit() and "@" not in v and not v.startswith("$"):
                        return v
    except:
        pass

    # Method 3: MaterialIdName -- split on "|" pipe to get name after library prefix
    # e.g. "Steel|Plain Carbon Steel" -> "Plain Carbon Steel"
    # Only use if result is not a pure number (which means it is an index)
    try:
        raw = str(model.MaterialIdName or "").strip()
        if raw:
            parts = raw.split("|")
            mat = parts[-1].strip()   # take last segment after all pipes
            if mat and not mat.isdigit():
                return mat
    except:
        pass

    # Method 4: GetModelDoc material via configuration-specific call
    try:
        cfg = model.ConfigurationManager.ActiveConfiguration.Name
        mat_db = model.GetMaterialPropertyName(cfg)
        if mat_db and str(mat_db).strip() and not str(mat_db).strip().isdigit():
            return str(mat_db).strip()
    except:
        pass

    return ""


def _resolve_sys_prop(model, sw_prop):
    """Resolve SW system property names via direct API calls."""
    name = sw_prop.strip().upper()
    try:
        if "MATERIAL" in name:
            return _get_material_name(model)
        if "MASS" in name:
            mp = model.Extension.CreateMassProperty
            if mp:
                mp.UseSystemUnits = True
                return str(round(float(mp.Mass or 0), 4))
        if "FILE" in name:
            return os.path.splitext(os.path.basename(model.GetPathName))[0]
    except:
        pass
    return ""


def _clean_val(val_eval, val_out, model):
    """
    Return the best resolved value from a Get5 result pair.
    val_eval = index 3 (evaluated), val_out = index 2 (raw formula).
    Never returns an expression string.
    """
    ve = str(val_eval or "").strip()
    vo = str(val_out  or "").strip()

    # Best case: evaluated value is clean
    if ve and not _is_expr(ve):
        return ve

    # Evaluated is still an expression -- resolve the SW system prop
    src = vo if vo else ve
    if _is_expr(src):
        sw_prop = ""
        if "@" in src:
            sw_prop = src.split("@")[0].strip()
        elif "$PRP" in src.upper():
            m = re.search(r'"([^"]+)"', src)
            if m:
                sw_prop = m.group(1).strip()
        if sw_prop:
            resolved = _resolve_sys_prop(model, sw_prop)
            if resolved and not _is_expr(resolved):
                return resolved

    return ""


# =====================================================
# BATCH PROPERTY READ -- ONE PASS PER COMPONENT
# Reads ALL custom properties in a single COM loop,
# caches by (path::cfg). get_cp() then does zero
# additional COM calls -- pure dict lookup.
# =====================================================
def _read_cpm_value(cpm, name, model):
    """
    Read one property value from an open CustomPropertyManager.
    Returns clean resolved string or empty string.
    """
    n = str(name).strip()

    # Method 1: Get5 -- (retval, type, valOut, valEval, wasResolved)
    try:
        res = cpm.Get5(n, False, "", "", False)
        if isinstance(res, (list, tuple)) and len(res) >= 4:
            ve = str(res[3] or "").strip()  # evaluated value
            vo = str(res[2] or "").strip()  # raw/formula value
            if ve and not _is_expr(ve):
                return ve
            if vo and not _is_expr(vo):
                return vo
            # Both are expressions -- resolve linked SW system property
            src = ve if ve else vo
            if _is_expr(src):
                sw_prop = ""
                if "@" in src:
                    sw_prop = src.split("@")[0].strip()
                elif "$PRP" in src.upper():
                    m = re.search(r'"([^"]+)"', src)
                    if m:
                        sw_prop = m.group(1).strip()
                if sw_prop:
                    resolved = _resolve_sys_prop(model, sw_prop)
                    if resolved and not _is_expr(resolved):
                        return resolved
    except:
        pass

    # Method 2: CustomInfo2
    try:
        v = str(model.CustomInfo2("", n) or "").strip()
        if v and not _is_expr(v):
            return v
    except:
        pass

    # Method 3: GetCustomInfoValue
    try:
        v = str(model.GetCustomInfoValue("", n) or "").strip()
        if v and not _is_expr(v):
            return v
    except:
        pass

    return ""


def _get_sw_special_props(model):
    """
    Read SW built-in system properties that never appear in GetNames.
    Uses _get_material_name() which tries multiple APIs to get the
    correct display name -- avoids the MaterialIdName index number bug.
    """
    special = {}
    try:
        mat = _get_material_name(model)
        if mat and mat.lower() not in ("", "unknown", "-", "none"):
            special["material"] = mat
    except:
        pass
    try:
        fname = os.path.splitext(os.path.basename(model.GetPathName))[0]
        if fname:
            special["sw-bom part number"] = fname
    except:
        pass
    return special


def _load_props(comp):
    """
    Read ALL properties for a component in one batch -- 3 passes:
    1. Config-specific custom properties via GetNames + Get5
    2. Global custom properties via GetNames + Get5
    3. SW special/system properties (Material from appearance, etc.)
    Plus a targeted Pass 4 for any expected CP still blank after passes 1-3.
    Result cached per (path::cfg).
    """
    try:
        model = comp.GetModelDoc2
        if model is None:
            return {}

        path = model.GetPathName
        cfg  = ""
        try:
            cfg = comp.ReferencedConfiguration
        except:
            pass

        key = path + "::" + cfg
        if key in _prop_cache:
            return _prop_cache[key]

        props = {}

        # Pass 1 + 2: user-defined custom properties
        for c in [cfg, ""]:
            try:
                cpm = model.Extension.CustomPropertyManager(c)
                if cpm is None:
                    continue
                names = cpm.GetNames
                if not names:
                    continue
                for n in names:
                    k = str(n).strip().lower()
                    if not k or k in props:
                        continue
                    v = _read_cpm_value(cpm, n, model)
                    if v:
                        props[k] = v
                        props[str(n).strip()] = v   # original-case copy
            except:
                continue

        # Pass 3: SW special properties to fill gaps
        for k, v in _get_sw_special_props(model).items():
            if k not in props:
                props[k] = v

        # Pass 4: targeted direct lookup for each expected CP still missing
        for cp_name in [CP_PART_CODE, CP_PART_NAME, CP_DESCRIPTION,
                        CP_REVISION, CP_MATERIAL, CP_THICKNESS,
                        CP_PROCESS, CP_PROCESS1, CP_PROCESS2,
                        CP_REQUIRED, CP_REMARK]:
            if cp_name.lower() in props:
                continue
            for c in [cfg, ""]:
                try:
                    cpm = model.Extension.CustomPropertyManager(c)
                    if cpm is None:
                        continue
                    v = _read_cpm_value(cpm, cp_name, model)
                    if v:
                        props[cp_name.lower()] = v
                        props[cp_name] = v
                        break
                except:
                    continue

        _prop_cache[key] = props
        mat_status = "material='" + props.get("material", "") + "'" if "material" in props else "material=BLANK"
        print(f"    [props] {os.path.basename(path)}: {len(props)} props | {mat_status}")
        return props
    except:
        return {}


def get_cp(comp, prop_name):
    """Single property lookup from cached batch. Zero extra COM calls."""
    props = _load_props(comp)
    return (props.get(prop_name.lower(), "")
            or props.get(prop_name, "")
            or props.get(prop_name.strip(), ""))


# =====================================================
# MASS -- CACHED
# =====================================================
def get_mass_kg(comp):
    """Mass in kg, cached per (path::cfg) -- SW recalculates only once per part."""
    try:
        model = comp.GetModelDoc2
        if model is None:
            return 0.0
        path = model.GetPathName
        cfg  = ""
        try:
            cfg = comp.ReferencedConfiguration
        except:
            pass
        key = path + "::" + cfg
        if key in _mass_cache:
            return _mass_cache[key]
        mp = model.Extension.CreateMassProperty
        if mp is None:
            _mass_cache[key] = 0.0
            return 0.0
        mp.UseSystemUnits = True
        mass = round(float(mp.Mass or 0), 4)
        _mass_cache[key] = mass
        return mass
    except:
        return 0.0


# =====================================================
# WALK ASSEMBLY TREE
# =====================================================
def walk_assembly(parent_comp, parent_hierarchy, level, rows, qty_multiplier=1):
    try:
        children = list(parent_comp.GetChildren or [])
    except:
        return

    grouped   = {}
    order_map = {}
    order_idx = 0

    for child in children:
        try:
            if child.IsSuppressed:
                continue
            model = child.GetModelDoc2
            if model is None:
                continue
            key = model.GetPathName + "::" + child.ReferencedConfiguration
            if key not in grouped:
                grouped[key]   = []
                order_map[key] = order_idx
                order_idx += 1
            grouped[key].append(child)
        except:
            continue

    sorted_keys = sorted(grouped.keys(), key=lambda k: order_map[k])

    child_counter = 0
    for key in sorted_keys:
        comp_list = grouped[key]
        child_counter += 1
        comp      = comp_list[0]
        unit_qty  = len(comp_list)
        total_qty = unit_qty * qty_multiplier
        hierarchy = (parent_hierarchy + "." if parent_hierarchy else "") + str(child_counter)

        model = comp.GetModelDoc2
        path  = model.GetPathName if model else ""
        fname = os.path.splitext(os.path.basename(path))[0] if path else ""

        # All properties -- single batch read per component (cached)
        part_code   = get_cp(comp, CP_PART_CODE)   or fname
        part_name   = get_cp(comp, CP_PART_NAME)   or fname
        description = get_cp(comp, CP_DESCRIPTION)
        revision    = get_cp(comp, CP_REVISION)
        material    = get_cp(comp, CP_MATERIAL)
        thickness   = get_cp(comp, CP_THICKNESS)
        process     = get_cp(comp, CP_PROCESS)
        process1    = get_cp(comp, CP_PROCESS1)
        process2    = get_cp(comp, CP_PROCESS2)
        required    = get_cp(comp, CP_REQUIRED)
        remark      = get_cp(comp, CP_REMARK)

        if not thickness:
            m = re.search(r'-T(\d+\.?\d*)(?:-|_|$)', part_code.upper())
            if m:
                thickness = m.group(1)

        unit_wt  = get_mass_kg(comp)
        total_wt = round(unit_wt * total_qty, 4)

        rows.append({
            "Hierarchy"        : hierarchy,
            "Level"            : level,
            "Part Code"        : part_code,
            "Part Name"        : part_name,
            "Description"      : description,
            "Revision"         : revision,
            "Material"         : material,
            "Unit Weight (kg)" : unit_wt,
            "Total Weight (kg)": total_wt,
            "Thickness (mm)"   : thickness,
            "Unit Qty"         : unit_qty,
            "Total Qty"        : total_qty,
            "Process"          : process,
            "Required"         : required,
            "Remark"           : remark,
            "Process 1"        : process1,
            "Process 2"        : process2,
        })

        cp_warn = " [CP blank]" if not description and not material else ""
        print(f"  {'  '*(level-1)}{hierarchy}  {part_code}  Qty:{unit_qty}{cp_warn}")

        try:
            if model and model.GetType == SW_PART_TYPE_ASSEMBLY:
                walk_assembly(comp, hierarchy, level + 1, rows, total_qty)
        except:
            pass


# =====================================================
# EXPORT TO EXCEL
# =====================================================
def export_to_excel(rows, assembly_name, output_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    COLS = [
        "Hierarchy", "Level", "Part Code", "Part Name", "Description",
        "Revision", "Material", "Unit Weight (kg)", "Total Weight (kg)",
        "Thickness (mm)", "Unit Qty", "Total Qty", "Process",
        "Required", "Remark", "Process 1", "Process 2"
    ]
    NAVY="1B2A4A"; TEAL="1E6E6E"; WHITE="FFFFFF"; ALT="EEF2F7"; LIGHT="F0F4F8"

    def fill(hx):
        return PatternFill(start_color=hx, end_color=hx, fill_type="solid")
    thin = Side(style='thin', color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"

    ws['A1'] = "  BOM Export -- " + assembly_name
    ws['A1'].font = Font(bold=True, size=13, color=WHITE, name="Montserrat")
    ws['A1'].fill = fill(NAVY)
    ws['A1'].alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 26
    for c in range(2, len(COLS)+1):
        ws.cell(row=1, column=c).fill = fill(NAVY)

    for ci, (lbl, val) in enumerate([
        ("Assembly",    assembly_name),
        ("Exported",    datetime.today().strftime("%d-%m-%Y %H:%M")),
        ("Total Parts", len(rows)),
    ]):
        lc = ws.cell(row=2, column=ci*2+1, value=lbl)
        vc = ws.cell(row=2, column=ci*2+2, value=val)
        lc.font=Font(bold=True,size=9,color=WHITE,name="Montserrat"); lc.fill=fill(TEAL)
        lc.alignment=Alignment(horizontal='right',vertical='center',indent=1); lc.border=bdr
        vc.font=Font(size=9,name="Montserrat"); vc.fill=fill(LIGHT)
        vc.alignment=Alignment(horizontal='left',vertical='center',indent=1); vc.border=bdr
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 5

    ws.row_dimensions[4].height = 22
    for ci, col in enumerate(COLS, start=1):
        c = ws.cell(row=4, column=ci, value=col)
        c.fill=fill(TEAL); c.font=Font(bold=True,size=9,color=WHITE,name="Montserrat")
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=bdr

    for ri, row in enumerate(rows, start=5):
        fhx = WHITE if ri%2==0 else ALT
        for ci, col in enumerate(COLS, start=1):
            c = ws.cell(row=ri, column=ci, value=row.get(col,""))
            c.fill=fill(fhx); c.font=Font(size=9,name="Montserrat")
            c.border=bdr; c.alignment=Alignment(vertical='center')

    wmap = {
        "Hierarchy":13,"Level":7,"Part Code":28,"Part Name":28,
        "Description":36,"Revision":10,"Material":16,
        "Unit Weight (kg)":16,"Total Weight (kg)":16,
        "Thickness (mm)":14,"Unit Qty":10,"Total Qty":10,
        "Process":14,"Required":10,"Remark":18,
        "Process 1":16,"Process 2":14,
    }
    for ci, col in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = wmap.get(col,14)

    ws.freeze_panes = ws.cell(row=5, column=1)
    wb.save(output_path)
    print(f"  Saved: {output_path}")


# =====================================================
# MAIN EXPORT
# =====================================================
def run_export(sldasm_path, output_dir, progress_bar, root_window, status_label):
    pythoncom.CoInitialize()
    try:
        def upd(msg, pct):
            status_label.config(text=msg)
            progress_bar["value"] = pct
            root_window.update()

        print("\n" + "="*60)
        print("  SW BOM EXPORTER  v3.0")
        print("="*60)

        # Clear caches for fresh export
        _prop_cache.clear()
        _mass_cache.clear()

        upd("Connecting to SolidWorks ...", 8)
        sw_app, err = connect_to_solidworks()
        model = None

        if sw_app is None:
            if not sldasm_path:
                messagebox.showerror("SolidWorks Not Running",
                    "SolidWorks is not running.\n\nPlease browse to a .SLDASM file.")
                upd("Ready", 0); return
            upd("Launching SolidWorks ...", 14)
            sw_app, model, err = launch_and_open(sldasm_path)
            if err:
                messagebox.showerror("Launch Failed", "Could not launch SolidWorks:\n\n"+err)
                upd("Error", 0); return
        else:
            upd("Getting active assembly ...", 15)
            if sldasm_path:
                errors   = win32.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
                warnings = win32.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
                model = sw_app.OpenDoc6(sldasm_path, SW_PART_TYPE_ASSEMBLY, 1, "", errors, warnings)
                if model is None:
                    messagebox.showerror("Open Failed",
                        "Could not open:\n"+sldasm_path+"\n\nSW Error: "+str(errors.value))
                    upd("Error", 0); return
            else:
                model, err = get_active_assembly(sw_app)
                if err:
                    messagebox.showerror("No Assembly", err)
                    upd("Error", 0); return

        assembly_name = os.path.splitext(os.path.basename(model.GetPathName))[0]
        print(f"\n  Assembly : {assembly_name}")

        upd("Reading assembly tree ...", 22)
        root_comp = model.ConfigurationManager.ActiveConfiguration.GetRootComponent3(True)
        if root_comp is None:
            messagebox.showerror("Error", "Could not get root component.")
            upd("Error", 0); return

        upd("Building BOM ...", 35)
        print("\n  Component tree:")
        rows = []
        walk_assembly(root_comp, "", 1, rows, qty_multiplier=1)

        if not rows:
            messagebox.showwarning("Empty BOM", "No components found.")
            upd("Done -- empty", 100); return

        print(f"\n  Total rows : {len(rows)}")
        print(f"  Props cached for {len(_prop_cache)} unique parts")
        print(f"  Mass cached for {len(_mass_cache)} unique parts")

        cp_filled = sum(1 for r in rows if r.get("Description") or r.get("Material"))
        if cp_filled == 0:
            print("\n  WARNING: All custom properties blank.")

        upd("Exporting to Excel ...", 72)
        if not output_dir:
            output_dir = os.path.dirname(model.GetPathName)
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        out = os.path.join(output_dir, f"{assembly_name}_BOM_{ts}.xlsx")
        export_to_excel(rows, assembly_name, out)

        upd("Done!", 100)
        print("="*60)
        print("  EXPORT COMPLETED!")
        print("="*60)

        cp_warn = ("\n\nWARNING: Custom properties all blank.\n"
                   "Check CP names match code constants."
                   if cp_filled == 0 else "")

        messagebox.showinfo("Success",
            "BOM exported!\n\n"
            "Assembly   : " + assembly_name + "\n"
            "Total rows : " + str(len(rows)) + "\n\n"
            "Saved as:\n  " + os.path.basename(out) + "\n\n"
            "Location:\n  " + output_dir + cp_warn)

    except Exception as e:
        progress_bar["value"] = 0
        status_label.config(text="Error -- check terminal")
        print(f"\nERROR: {e}")
        import traceback; traceback.print_exc()
        messagebox.showerror("Error", "An error occurred:\n\n" + str(e))
    finally:
        pythoncom.CoUninitialize()

_DEV = ''.join(['M','a','h','e','s','h',' ','A','r','v','i','n','d',' ','C','h','a','v','a','n'])

# =====================================================
# GUI
# =====================================================
def create_gui():
    NAVY="#1B2A4A"; TEAL="#1E6E6E"; ACCENT="#E8A838"
    BG="#F0F4F8"; WHITE="#FFFFFF"; GREY="#7A8A99"; DARK="#2C3E50"

    root = tk.Tk()
    root.title("SW BOM Exporter  v3.0")
    root.configure(bg=BG)
    root.geometry("720x590")
    root.resizable(False, False)
    root.lift()
    root.focus_force()

    hdr = tk.Frame(root, bg=NAVY, height=68)
    hdr.pack(fill=tk.X)
    hdr.pack_propagate(False)
    tk.Label(hdr, text="SolidWorks BOM Exporter",
             font=("Montserrat",16,"bold"), bg=NAVY, fg=WHITE).pack(side=tk.LEFT, padx=22, pady=14)
    tk.Label(hdr, text="v3.0",
             font=("Montserrat",10), bg=NAVY, fg=ACCENT).pack(side=tk.LEFT, pady=20)

    ftr = tk.Frame(root, bg=NAVY, height=28)
    ftr.pack(fill=tk.X, side=tk.BOTTOM)
    ftr.pack_propagate(False)
    tk.Label(ftr,
             text="Concept & Developed by "+_DEV+"  |  SW BOM Exporter v3.0  |  NIDO Automation",
             bg=NAVY, fg=GREY, font=("Montserrat",8)).pack(pady=6)

    body = tk.Frame(root, bg=BG)
    body.pack(fill=tk.BOTH, expand=True, padx=22, pady=14)

    sf = tk.LabelFrame(body, text="  How it works  ", bg=BG, fg=TEAL,
                       font=("Montserrat",9,"bold"), bd=1, relief=tk.GROOVE, padx=8, pady=6)
    sf.pack(fill=tk.X, pady=(0,10))
    for num, text in [
        ("1", "Connects to running SolidWorks automatically; browse .SLDASM if not open"),
        ("2", "Walks full indented BOM tree recursively through all levels"),
        ("3", "v3.0 -- batch property read: all CPs fetched in ONE pass per part (fast)"),
        ("4", "Mass cached per unique part -- SW recalculates only once, not per instance"),
        ("5", "Resolves SW link expressions like SW-Material@FileName automatically"),
        ("6", "Exports Excel matching Nido_Mfg_BOM_R3 template column format exactly"),
    ]:
        rf = tk.Frame(sf, bg=BG)
        rf.pack(fill=tk.X, pady=1)
        bn = TEAL if num.isdigit() else BG
        fn = WHITE if num.isdigit() else BG
        tk.Label(rf, text=" "+num+" ", bg=bn, fg=fn,
                 font=("Montserrat",8,"bold")).pack(side=tk.LEFT, padx=(0,7))
        tk.Label(rf, text=text, bg=BG, fg=DARK,
                 font=("Montserrat",9), anchor='w').pack(side=tk.LEFT)

    def make_row(label, browse_fn, bc=None):
        lf = tk.LabelFrame(body, text="  "+label+"  ", bg=BG, fg=TEAL,
                            font=("Montserrat",9,"bold"), bd=1, relief=tk.GROOVE, padx=10, pady=8)
        lf.pack(fill=tk.X, pady=(0,8))
        fi = tk.Frame(lf, bg=BG); fi.pack(fill=tk.X)
        entry = tk.Entry(fi, width=62, font=("Montserrat",9), bg=WHITE,
                         relief=tk.FLAT, highlightthickness=1, highlightbackground="#CCCCCC")
        entry.pack(side=tk.LEFT, padx=(0,8), ipady=5)
        bc2 = bc or ACCENT
        fc  = NAVY if bc2==ACCENT else WHITE
        btn = tk.Button(fi, text="Browse...", command=lambda e=entry: browse_fn(e),
                        bg=bc2, fg=fc, font=("Montserrat",9,"bold"),
                        relief=tk.FLAT, cursor="hand2", padx=10, pady=5)
        btn.pack(side=tk.LEFT)
        if bc2 == ACCENT:
            btn.bind("<Enter>", lambda e,b=btn: b.config(bg="#D4932A"))
            btn.bind("<Leave>", lambda e,b=btn: b.config(bg=ACCENT))
        return entry

    def br_asm(e):
        fn = filedialog.askopenfilename(title="Select Assembly",
                                        filetypes=[("SolidWorks Assembly","*.SLDASM *.sldasm"),("All","*.*")])
        if fn: e.delete(0,tk.END); e.insert(0,fn)

    def br_out(e):
        fn = filedialog.askdirectory(title="Select Output Folder")
        if fn: e.delete(0,tk.END); e.insert(0,fn)

    asm_entry = make_row("Assembly File (.SLDASM) -- optional if SolidWorks is already open", br_asm)
    out_entry = make_row("Output Folder -- leave blank to save next to assembly", br_out, bc="#7A8A99")

    pgf = tk.Frame(body, bg=BG); pgf.pack(fill=tk.X, pady=(6,0))
    style = ttk.Style(); style.theme_use('clam')
    style.configure("SW.Horizontal.TProgressbar", troughcolor="#D0D8E4", background=TEAL, thickness=14)
    pb = ttk.Progressbar(pgf, length=676, mode='determinate', style="SW.Horizontal.TProgressbar")
    pb.pack(pady=(6,3))

    sl = tk.Label(body, text="Ready -- press EXPORT BOM to begin",
                  bg=BG, fg=GREY, font=("Montserrat",9,"italic"))
    sl.pack(pady=(3,8))

    def run_diagnostics():
        sl.config(text="Running diagnostics ...", fg=TEAL); root.update()
        def _d():
            pythoncom.CoInitialize()
            try:
                sw, _ = connect_to_solidworks()
                if sw is None:
                    messagebox.showerror("Not Connected","SolidWorks not running."); return
                doc, err = get_active_assembly(sw)
                if err:
                    messagebox.showerror("No Assembly", err); return
                rc = doc.ConfigurationManager.ActiveConfiguration.GetRootComponent3(True)
                kids = list(rc.GetChildren or [])
                test = next((c for c in kids if not c.IsSuppressed), None)
                if test is None:
                    print("  No non-suppressed children."); return
                m = test.GetModelDoc2
                fname = os.path.splitext(os.path.basename(m.GetPathName if m else ""))[0]
                print(f"\n  DIAGNOSTIC -- {fname}")
                # Load all props via batch
                props = _load_props(test)
                print(f"  Total custom properties found: {len(props)}")
                for k, v in props.items():
                    print(f"    '{k}' = '{v}'")
                print("\n  Checking expected CP names:")
                for cp in [CP_PART_CODE, CP_PART_NAME, CP_DESCRIPTION,
                           CP_REVISION, CP_MATERIAL, CP_THICKNESS,
                           CP_PROCESS, CP_PROCESS1, CP_PROCESS2, CP_REQUIRED, CP_REMARK]:
                    val = get_cp(test, cp)
                    print(f"    {cp:<22} -> {repr(val) if val else '[NOT FOUND]'}")
                messagebox.showinfo("Diagnostics Done",
                    "Check terminal for results.\n\n"
                    "All found property names and values are listed.")
                sl.config(text="Diagnostics done -- check terminal", fg=TEAL)
            except Exception as ex:
                print(f"  Diag error: {ex}")
                import traceback; traceback.print_exc()
            finally:
                pythoncom.CoUninitialize()
        threading.Thread(target=_d, daemon=True).start()

    def start():
        fp = asm_entry.get().strip()
        od = out_entry.get().strip()
        if od and not os.path.isdir(od):
            messagebox.showerror("Invalid Folder", "Output folder not found:\n"+od); return
        if fp and not os.path.exists(fp):
            messagebox.showerror("Not Found", "Assembly file not found:\n"+fp); return
        pb["value"] = 0
        sl.config(text="Connecting to SolidWorks ...", fg=TEAL); root.update()
        threading.Thread(target=run_export, args=(fp, od, pb, root, sl), daemon=True).start()

    btn_row = tk.Frame(body, bg=BG); btn_row.pack(pady=2)

    sb = tk.Button(btn_row, text="EXPORT BOM", command=start,
                   bg=TEAL, fg=WHITE, font=("Montserrat",12,"bold"),
                   relief=tk.FLAT, cursor="hand2", padx=44, pady=10)
    sb.pack(side=tk.LEFT, padx=(0,10))
    sb.bind("<Enter>", lambda e: sb.config(bg=NAVY))
    sb.bind("<Leave>", lambda e: sb.config(bg=TEAL))

    db = tk.Button(btn_row, text="DIAGNOSTICS", command=run_diagnostics,
                   bg="#7A8A99", fg=WHITE, font=("Montserrat",10,"bold"),
                   relief=tk.FLAT, cursor="hand2", padx=16, pady=10)
    db.pack(side=tk.LEFT)
    db.bind("<Enter>", lambda e: db.config(bg=DARK))
    db.bind("<Leave>", lambda e: db.config(bg="#7A8A99"))

    root.mainloop()


if __name__ == "__main__":
    print("\n" + "="*60)
    print("  INITIALIZING SW BOM EXPORTER  v3.0")
    print("="*60)
    create_gui()